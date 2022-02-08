# -*- coding: utf-8 -*-
"""
Created on Mon Jan 25 10:07:19 2021

@author: linjianing
"""

import os
import math
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from sklearn.linear_model import LogisticRegression
from sklearn.model_selection import GridSearchCV
import statsmodels.api as sm
from .stepwise import stepwise_selection


class LassoLRCV:
    """lasso交叉验证逻辑回归."""

    def __init__(self, entity):
        self.entity = entity

    def fit(self, verbose=False):
        """训练."""
        entity = self.entity
        y_ser = entity.pipe_y.copy(deep=True)
        X_df = entity.pipe_X.copy(deep=True)
        X_names = X_df.columns.to_list()
        params = {'C': 1/np.logspace(np.log(1e-6), np.log(1), 50, base=math.e)}
        lass_lr = LogisticRegression(penalty='l1', solver='liblinear')
        while True:
            gscv = GridSearchCV(lass_lr, params)
            gscv.fit(X_df, y_ser)
            if not any(gscv.best_estimator_.coef_.ravel() < 0):
                break
            X_names = [
                k for k, v in
                dict(zip(X_names, gscv.best_estimator_.coef_.ravel())).items()
                if v > 0]
            if verbose:
                print(X_names)
            X_df = X_df.loc[:, X_names]
        coef_dict = dict(zip(X_names, gscv.best_estimator_.coef_.ravel()))
        self.lasso_vars = [k for k, v in coef_dict.items() if v > 0]
        entity.steps.update({'lasso': self.lasso_vars})
        return self

    def transform(self, entity=None):
        """应用."""
        if entity is None:
            entity = self.entity
        drop_vars = entity.pipe_X.columns.difference(self.lasso_vars)
        entity.pipe_X = entity.pipe_X.drop(drop_vars, axis=1)
        entity.update_data()
        return self

    def output(self):
        """输出报告."""
        pass

    def report(self, file):
        df = pd.DataFrame({'variable': self.lasso_vars})
        if not os.path.exists(file):
            with pd.ExcelWriter(file) as writer:
                df.to_excel(writer, sheet_name='lassoLR')
        else:
            with pd.ExcelWriter(file, engine='openpyxl') as writer:
                book = load_workbook(file)
                writer.book = book
                df.to_excel(writer, sheet_name='lassoLR')
        return self


class Stepwise:
    """逐步回归筛选."""

    def __init__(self, variable_options={}, threshold_in=0.01,
                 threshold_out=0.05):
        self.threshold_in = threshold_in
        self.threshold_out = threshold_out

    def fit(self, X, y, verbose=True):
        """训练."""
        X_df = X.copy(deep=True)
        step_out = stepwise_selection(
            X_df, y, threshold_in=self.threshold_in,
            threshold_out=self.threshold_out, verbose=verbose)
        self.stepwise_vars = step_out
        return self

    def transform(self, X):
        """应用."""
        return X.loc[:, self.stepwise_vars].copy(deep=True)

    def output(self):
        """输出报告."""
        pass


class LRCust:
    """逻辑回归."""

    def __init__(self, entity):
        self.entity = entity

    def fit(self):
        """训练."""
        entity = self.entity
        y_ser = entity.pipe_y.copy(deep=True)
        X_df = entity.pipe_X.copy(deep=True)
        lr = sm.Logit(y_ser, sm.add_constant(X_df)).fit()
        self.LRmodel_ = lr
        self.LRCust_vars = lr.summary2().tables[1].loc[:, 'Coef.'].index\
            .difference(['const'])
        coef = self.LRmodel_.summary2().tables[1]
        coef.loc[:, 'weight'] = X_df.apply(lambda x: x.max()-x.min())\
            * coef.loc[:, 'Coef.']
        coef.loc[:, 'weight'] = coef.loc[:, 'weight']\
            / coef.loc[:, 'weight'].sum() * 100
        self.coef = coef
        entity.steps.update({'LRCust': self.coef.to_dict(orient='index')})
        return self

    def transform(self, entity=None):
        """应用."""
        if entity is None:
            entity = self.entity
        drop_vars = entity.pipe_X.columns.difference(self.LRCust_vars)
        entity.pipe_X = entity.pipe_X.drop(drop_vars, axis=1)
        entity.update_data()
        self.predict()
        return self

    def predict(self):
        """应用."""
        entity = self.entity
        X_df = entity.pipe_X.loc[:, self.LRCust_vars].copy(deep=True)
        lr = self.LRmodel_
        trns = lr.predict(sm.add_constant(X_df))
        trns.name = 'proba'
        entity.proba = trns
        return self

    def output(self):
        """输出报告."""
        return self.coef

    def report(self, file):
        df = self.coef
        if not os.path.exists(file):
            with pd.ExcelWriter(file) as writer:
                df.to_excel(writer, sheet_name='LRCust')
        else:
            with pd.ExcelWriter(file, engine='openpyxl') as writer:
                book = load_workbook(file)
                writer.book = book
                df.to_excel(writer, sheet_name='LRCust')
        return self


# class LassoLRCV:
#     """lasso交叉验证逻辑回归."""

#     def __init__(self, variable_options={}):
#         pass

#     def fit(self, X, y, verbose=False):
#         """训练."""
#         X_df = X.copy(deep=True)
#         y_ser = y.copy(deep=True)
#         X_names = X_df.columns.to_list()
#         params = {'C': 1/np.logspace(np.log(1e-6), np.log(1), 50, base=math.e)}
#         lass_lr = LogisticRegression(penalty='l1', solver='liblinear')
#         while True:
#             gscv = GridSearchCV(lass_lr, params)
#             gscv.fit(X_df, y_ser)
#             if not any(gscv.best_estimator_.coef_.ravel() < 0):
#                 break
#             X_names = [k for k, v in dict(zip(X_names, gscv.best_estimator_.coef_.ravel())).items() if v > 0]
#             if verbose:
#                 print(X_names)
#             X_df = X_df.loc[:, X_names]
#         coef_dict = dict(zip(X_names, gscv.best_estimator_.coef_.ravel()))
#         self.lasso_vars = [k for k, v in coef_dict.items() if v > 0]
#         return self

#     def transform(self, X):
#         """应用."""
#         return X.loc[:, self.lasso_vars].copy(deep=True)

#     def output(self):
#         """输出报告."""
#         pass


# class Stepwise:
#     """逐步回归筛选."""

#     def __init__(self, variable_options={}, threshold_in=0.01,
#                  threshold_out=0.05):
#         self.threshold_in = threshold_in
#         self.threshold_out = threshold_out

#     def fit(self, X, y, verbose=True):
#         """训练."""
#         X_df = X.copy(deep=True)
#         step_out = stepwise_selection(
#             X_df, y, threshold_in=self.threshold_in,
#             threshold_out=self.threshold_out, verbose=verbose)
#         self.stepwise_vars = step_out
#         return self

#     def transform(self, X):
#         """应用."""
#         return X.loc[:, self.stepwise_vars].copy(deep=True)

#     def output(self):
#         """输出报告."""
#         pass


# class LRCust:
#     """逻辑回归."""

#     def __init__(self, variable_options={}):
#         pass

#     def fit(self, X, y):
#         """训练."""
#         X_df = X.copy(deep=True)
#         y_ser = y.copy(deep=True)
#         lr = sm.Logit(y_ser, sm.add_constant(X_df)).fit()
#         self.LRmodel_ = lr
#         self.LRCust_vars = lr.summary2().tables[1].loc[:, 'Coef.'].index\
#             .difference(['const'])
#         coef = self.LRmodel_.summary2().tables[1]
#         coef.loc[:, 'weight'] = X_df.apply(lambda x: x.max()-x.min())\
#             * coef.loc[:, 'Coef.']
#         coef.loc[:, 'weight'] = coef.loc[:, 'weight']\
#             / coef.loc[:, 'weight'].sum() * 100
#         self.coef = coef
#         return self

#     def transform(self, X):
#         """应用."""
#         return X.copy(deep=True).loc[:, self.LRCust_vars]

#     def predict(self, X):
#         """应用."""
#         X_df = X.copy(deep=True).loc[:, self.LRCust_vars]
#         lr = self.LRmodel_
#         trns = lr.predict(sm.add_constant(X_df))
#         trns.name = 'proba'
#         return trns

#     def output(self):
#         """输出报告."""
#         return self.coef
